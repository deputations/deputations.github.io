"""Convert the master Ministries & Organisations xlsx -> data/ministries.json."""
import json
from collections import OrderedDict
from pathlib import Path
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
SRC = Path("H:/My Drive/Deputation/Master list - Ministries and Departments.xlsx")
OUT = REPO / "data" / "ministries.json"


def main():
    wb = load_workbook(SRC, data_only=True)
    ws = wb["Main"]

    grouped: "OrderedDict[str, list[dict]]" = OrderedDict()
    seen_pairs: set[tuple[str, str]] = set()

    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        # Columns: S. No. | Ministry | Organisation Name | Organisation Type
        if not row or len(row) < 4:
            continue
        ministry = (row[1] or "").strip() if isinstance(row[1], str) else ""
        org = (row[2] or "").strip() if isinstance(row[2], str) else ""
        org_type = (row[3] or "").strip() if isinstance(row[3], str) else ""
        if not ministry or not org:
            continue
        key = (ministry, org)
        if key in seen_pairs:
            continue
        seen_pairs.add(key)
        grouped.setdefault(ministry, []).append({"name": org, "type": org_type})

    payload = {
        "_generated": True,
        "_source": SRC.name,
        "ministries": [
            {
                "name": m,
                "organisations": sorted(orgs, key=lambda x: x["name"].lower()),
            }
            for m, orgs in sorted(grouped.items(), key=lambda kv: kv[0].lower())
        ],
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    total_orgs = sum(len(m["organisations"]) for m in payload["ministries"])
    print(f"wrote {OUT.relative_to(REPO)} — {len(payload['ministries'])} ministries, {total_orgs} organisations")


if __name__ == "__main__":
    main()
