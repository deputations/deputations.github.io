"""Build DEX data files from the anonymised deputation xlsx + master ministries.json.

Outputs (under data/dex/):
  organisations.json   — master org list, joined with DEX signal where available
  reports.json         — anonymised community records (no Suspect_Bribe ever)
  scores.json          — derived DEX score + confidence per org (regenerate, don't edit)
  aliases.json         — name → org-id lookup helpers (ministry + dept aliases)
  methodology.json     — versioned formula, weights, bands, changelog
  updates.json         — last-refresh metadata

The Allow_Suspect_Bribe column is read ONLY to set an internal moderation flag that
lowers confidence for the affected org. It is stripped from every public JSON.
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter, OrderedDict
from datetime import datetime, timezone
from pathlib import Path

import os

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
MINISTRIES_JSON = REPO / "data" / "ministries.json"
OUT_DIR = REPO / "data" / "defex"

# Source resolution order:
#   1. $DEX_GOOGLE_SHEET_ID  → live Google Sheet via Workload Identity (used by CI)
#   2. $DEX_SRC_XLSX         → explicit local path
#   3. data/raw/anonymised-deputation-data.xlsx  (committed snapshot)
#   4. local Downloads xlsx (dev fallback)
DEX_GOOGLE_SHEET_ID = (os.environ.get("DEX_GOOGLE_SHEET_ID") or "").strip()
_XLSX_CANDIDATES = [
    os.environ.get("DEX_SRC_XLSX"),
    str(REPO / "data" / "raw" / "anonymised-deputation-data.xlsx"),
    r"C:/Users/vivek/Downloads/anonymised Deputation data.xlsx",
]
SRC_XLSX = next((Path(p) for p in _XLSX_CANDIDATES if p and Path(p).exists()), None)

METHODOLOGY_VERSION = "0.2.0-beta"
SURVEY_URL = "https://script.google.com/macros/s/AKfycbzr1Jr7kFctZ4zHObJlPa8M_pR6zngWqx2dxClDFXkZ-QZSKa2fLVwtXTEca2lGICLB/exec"

EXPECTED_COLUMNS = [
    "Ministry", "Department", "Deputation_Allowed",
    "Allow_Vacancy_Based", "Allow_FCFS", "Allow_Senior_Only",
    "Allow_Highly_Beneficial", "Allow_Suspect_Bribe",
    "Conditional_Details", "Source_OM", "Source_Personal",
]


# ---------- input sources ----------------------------------------------------

def fetch_sheet_rows(sheet_id: str) -> tuple[list[dict], str]:
    """Pull rows from the first tab of a Google Sheet using ambient credentials.
    Returns (rows-as-dicts, source-description)."""
    from google.auth import default
    from googleapiclient.discovery import build as gapi_build

    credentials, _ = default(scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"])
    service = gapi_build("sheets", "v4", credentials=credentials, cache_discovery=False)

    meta = service.spreadsheets().get(spreadsheetId=sheet_id).execute()
    sheets = meta.get("sheets", [])
    if not sheets:
        raise RuntimeError("DEX sheet has no tabs.")
    title = sheets[0]["properties"]["title"]
    values = (service.spreadsheets().values()
              .get(spreadsheetId=sheet_id, range=f"'{title}'!A:ZZ")
              .execute()).get("values", [])
    if not values:
        raise RuntimeError("DEX sheet is empty.")

    headers = [str(h).strip() for h in values[0]]
    missing = [c for c in ("Ministry", "Department", "Deputation_Allowed") if c not in headers]
    if missing:
        raise RuntimeError(f"DEX sheet missing required columns: {missing}")

    rows = []
    for raw in values[1:]:
        padded = list(raw) + [""] * (len(headers) - len(raw))
        rec = {headers[i]: (str(padded[i]).strip() if padded[i] is not None else "") for i in range(len(headers))}
        rows.append(rec)
    return rows, f"google-sheet:{sheet_id[:8]}…/{title}"


def fetch_xlsx_rows(path: Path) -> tuple[list[dict], str]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header_row = list(next(it))
    headers = [h for h in header_row if h]
    rows = []
    for raw in it:
        if not raw or all(c is None or (isinstance(c, str) and not c.strip()) for c in raw):
            continue
        rec = {headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers))}
        rows.append(rec)
    return rows, f"xlsx:{path.name}"


def load_rows() -> tuple[list[dict], str]:
    if DEX_GOOGLE_SHEET_ID:
        return fetch_sheet_rows(DEX_GOOGLE_SHEET_ID)
    if SRC_XLSX is None:
        raise SystemExit(
            "No DEX source available. Set $DEX_GOOGLE_SHEET_ID, $DEX_SRC_XLSX, "
            "or place data/raw/anonymised-deputation-data.xlsx."
        )
    return fetch_xlsx_rows(SRC_XLSX)


# ---------- helpers ----------------------------------------------------------

def slugify(s: str) -> str:
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s or "x"


def normkey(s: str) -> str:
    """Aggressive normalisation for fuzzy matching: lowercase, strip punctuation,
    drop common stopwords and 'department/ministry/of'."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().lower()
    s = re.sub(r"\(.*?\)", " ", s)            # drop parenthetical e.g. (Secretariat)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    drop = {"ministry", "department", "dept", "of", "the", "and", "&", "for", "secretariat", "govt", "government", "india"}
    toks = [t for t in s.split() if t and t not in drop]
    return " ".join(toks)


def yn(v) -> bool:
    return isinstance(v, str) and v.strip().upper() == "YES"


def clamp(x, lo, hi):
    return max(lo, min(hi, x))


# ---------- master list ------------------------------------------------------

def load_master():
    data = json.loads(MINISTRIES_JSON.read_text(encoding="utf-8"))
    orgs = []
    ministry_keys = {}
    for m in data["ministries"]:
        mname = m["name"]
        ministry_keys[normkey(mname)] = mname
        for o in m["organisations"]:
            oname = o["name"]
            otype = o.get("type") or ""
            # Normalise org types the master mislabels (mostly "Attached Office"):
            #   "Ministry of X (Secretariat)" -> Ministry   (the ministry HQ)
            #   "Department of X"              -> Department
            if oname.strip().lower().endswith("(secretariat)"):
                otype = "Ministry"
            elif oname.strip().lower().startswith("department of"):
                otype = "Department"
            orgs.append({
                "id": f"{slugify(mname)}__{slugify(oname)}",
                "name": oname,
                "ministry": mname,
                "type": otype,
                "_nk_name": normkey(oname),
                "_nk_ministry": normkey(mname),
            })
    return orgs, ministry_keys


# ---------- matching ---------------------------------------------------------

def match_org(orgs, ministry: str, department: str):
    nk_min = normkey(ministry)
    nk_dep = normkey(department)
    if not nk_dep:
        return None, "empty-department"

    pool = [o for o in orgs if o["_nk_ministry"] == nk_min] or orgs

    # exact normalised match
    for o in pool:
        if o["_nk_name"] == nk_dep:
            return o, "exact"

    # secretariat handling: "ministry of X (secretariat)" -> match ministry-named org
    if "secretariat" in (department or "").lower():
        for o in pool:
            if o["_nk_name"] == normkey(ministry):
                return o, "secretariat"
            if o["_nk_name"].startswith("department " + nk_min.split()[0] if nk_min else ""):
                return o, "secretariat-loose"

    # substring either direction
    for o in pool:
        if nk_dep and (nk_dep in o["_nk_name"] or o["_nk_name"] in nk_dep):
            return o, "substring"

    # token overlap >= 2
    dep_tokens = set(nk_dep.split())
    best = None
    best_score = 0
    for o in pool:
        ot = set(o["_nk_name"].split())
        overlap = len(dep_tokens & ot)
        if overlap > best_score:
            best_score = overlap
            best = o
    if best and best_score >= 2:
        return best, f"tokens-{best_score}"

    return None, "no-match"


# ---------- scoring ----------------------------------------------------------

BASE_SCORE = {
    "Yes": 88,
    "Allow for some": 70,
    "Allowed with conditions": 58,
    "No": 18,
}

PENALTY_VACANCY = 18
PENALTY_SENIOR_ONLY = 15
PENALTY_FCFS = 8


def band_for(dex):
    if dex >= 80: return "Deputation-friendly"
    if dex >= 65: return "Generally supportive"
    if dex >= 50: return "Mixed"
    if dex >= 35: return "Restrictive"
    return "Very restrictive"


def _has_vacancy(row):
    raw = (row.get("Allow_Vacancy_Based") or "").strip()
    return yn(raw) or "vacancy" in raw.lower()


def aggregate_org(rows):
    """Combine ALL of an org's response rows into one cumulative score + confidence.

    Cumulative model — an org's ranking reflects every response collected for it,
    not a single row (the survey accumulates responses over the month):
      - Base B:    consensus Deputation_Allowed (most common value; ties -> most
                   conservative, i.e. lowest base score).
      - Penalties: a friction condition (vacancy / senior-only / FCFS) counts only
                   when a MAJORITY of responses report it — one outlier can neither
                   tank nor clear an org.
      - Evidence:  +8 if any official OM; +2 per community response, capped at +8.
      - Confidence rises with response volume and agreement, using the published
                   weights: reports .25 · official .40 · recency .15 · consistency
                   .10 · stage .10.
    Returns a score dict, or None if no response carries a usable Deputation_Allowed.
    Backward-compatible: a single row reproduces the old per-row score exactly.
    """
    n = len(rows)

    # Base — consensus Deputation_Allowed across the responses.
    valid = [(r.get("Deputation_Allowed") or "").strip() for r in rows]
    valid = [a for a in valid if a in BASE_SCORE]
    if not valid:
        return None
    counts = Counter(valid)
    top = max(counts.values())
    consensus = min((a for a in counts if counts[a] == top), key=lambda a: BASE_SCORE[a])
    B = BASE_SCORE[consensus]
    agreement = counts[consensus] / len(valid)            # 1.0 = unanimous

    # Penalties — applied only on majority report.
    def share(pred):
        return sum(1 for r in rows if pred(r)) / n
    penalties, signals = 0, []
    if share(_has_vacancy) >= 0.5:
        penalties += PENALTY_VACANCY
        signals.append({"key": "vacancy_based", "label": "Vacancy-based release", "weight": -PENALTY_VACANCY})
    if share(lambda r: yn(r.get("Allow_Senior_Only"))) >= 0.5:
        penalties += PENALTY_SENIOR_ONLY
        signals.append({"key": "senior_only", "label": "Senior-rank-only release", "weight": -PENALTY_SENIOR_ONLY})
    if share(lambda r: yn(r.get("Allow_FCFS"))) >= 0.5:
        penalties += PENALTY_FCFS
        signals.append({"key": "fcfs", "label": "FCFS / quota cap", "weight": -PENALTY_FCFS})
    if share(lambda r: yn(r.get("Allow_Highly_Beneficial"))) >= 0.5:
        signals.append({"key": "highly_beneficial", "label": "Release only if highly beneficial to parent", "weight": 0})

    # Evidence — official OM dominates; community reports accumulate (capped).
    n_om = sum(1 for r in rows if (r.get("Source_OM") or "").strip())
    n_personal = sum(1 for r in rows if (r.get("Source_Personal") or "").strip())
    has_om, has_personal = n_om > 0, n_personal > 0
    E = min((8 if has_om else 0) + 2 * n_personal, 8)

    dex = clamp(B - penalties + E, 0, 100)
    band = band_for(dex)
    if penalties > 0 and dex > 75:            # friction caps out of the "friendly" band
        dex, band = 75, "Generally supportive"

    # Confidence — methodology weights, now scaled by response volume + agreement.
    reports_cov = 1.0 if has_om else min(1.0, 0.5 + 0.5 * max(0, n_personal - 1) / 3)  # 1 report .5 … 4 .1.0
    c = (0.25 * reports_cov
         + 0.40 * (1.0 if has_om else 0.0)
         + 0.15 * 1.0
         + 0.10 * agreement
         + 0.10 * (0.4 if has_personal else 0.2))

    needs_review = any(yn(r.get("Allow_Suspect_Bribe")) for r in rows)
    if needs_review:
        c = min(c, 0.45)                      # cap at Medium until human-reviewed

    if c >= 0.70: conf_band = "High"
    elif c >= 0.45: conf_band = "Medium"
    elif c >= 0.20: conf_band = "Low"
    else: conf_band = "Insufficient"

    return {
        "dex": dex, "band": band, "signals": signals,
        "confidence": round(c, 3), "confidence_band": conf_band,
        "has_om": has_om, "has_personal": has_personal, "needs_review": needs_review,
    }


# ---------- main -------------------------------------------------------------

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    orgs, _ministry_keys = load_master()
    org_by_id = {o["id"]: o for o in orgs}

    rows_in, source_label = load_rows()

    reports = []
    score_map = {}
    unresolved = []
    synth_secretariats = {}
    org_rows = {}          # org_id -> [source rows], for cumulative scoring
    org_report_ids = {}    # org_id -> [report_id]

    for row in rows_in:
        ministry = (str(row.get("Ministry") or "")).strip()
        dept = (str(row.get("Department") or "")).strip()
        if not ministry and not dept:
            continue

        target, how = match_org(orgs, ministry, dept)

        # Fallback for ministry-level records (empty dept or "(Secretariat)").
        # First try matching against the existing master-list "Ministry X (Secretariat)"
        # entry; only synthesise a new org if that doesn't exist either.
        if not target and ministry and (not dept or "secretariat" in dept.lower()):
            sec_name_guess = f"{ministry} (Secretariat)"
            target, how = match_org(orgs, ministry, sec_name_guess)
            if not target:
                key = ministry
                if key not in synth_secretariats:
                    sec_id = f"{slugify(ministry)}__{slugify(sec_name_guess)}"
                    sec = {
                        "id": sec_id, "name": sec_name_guess, "ministry": ministry,
                        "type": "Ministry",
                        "_nk_name": normkey(sec_name_guess), "_nk_ministry": normkey(ministry),
                    }
                    orgs.append(sec)
                    org_by_id[sec_id] = sec
                    synth_secretariats[key] = sec
                target = synth_secretariats[key]
                how = "synth-secretariat"
            else:
                how = "secretariat-fallback"

        if not target:
            unresolved.append({"ministry": ministry, "department": dept, "reason": how})
            continue

        # Build the public report (Suspect_Bribe never leaks here) and stash the row;
        # scoring runs after the loop so each org's score reflects ALL its responses.
        report_id = f"r_{len(reports)+1:04d}"
        reports.append({
            "id": report_id,
            "org_id": target["id"],
            "type": "policy",  # survey will add 'experience'-type rows over time
            "deputation_allowed": (row.get("Deputation_Allowed") or "").strip() or None,
            "conditions": {
                "vacancy_based": (row.get("Allow_Vacancy_Based") or "").strip() or None,
                "fcfs": (row.get("Allow_FCFS") or "").strip() or None,
                "senior_only": yn(row.get("Allow_Senior_Only")),
                "highly_beneficial": yn(row.get("Allow_Highly_Beneficial")),
                "details": (row.get("Conditional_Details") or "").strip() or None,
            },
            "sources": {
                "om": (row.get("Source_OM") or "").strip() or None,
                "personal": (row.get("Source_Personal") or "").strip() or None,
            },
            "match_quality": how,
        })
        org_rows.setdefault(target["id"], []).append(row)
        org_report_ids.setdefault(target["id"], []).append(report_id)

    # ---------- cumulative scoring: one score per org, from ALL its responses ---
    for org_id, rows in org_rows.items():
        agg = aggregate_org(rows)
        if agg is None:
            continue
        score_map[org_id] = {**agg, "org_id": org_id, "report_ids": org_report_ids[org_id]}

    # ---------- organisations.json (public) ---------------------------------
    pub_orgs = []
    for o in orgs:
        s = score_map.get(o["id"])
        pub_orgs.append({
            "id": o["id"],
            "name": o["name"],
            "ministry": o["ministry"],
            "type": o["type"],
            "rated": s is not None,
            "dex": s["dex"] if s else None,
            "band": s["band"] if s else "Unrated",
            "confidence": s["confidence"] if s else 0,
            "confidence_band": s["confidence_band"] if s else "Insufficient",
            "reports": len(s["report_ids"]) if s else 0,
            "has_om": s["has_om"] if s else False,
        })

    # ---------- scores.json (public, strip needs_review) --------------------
    pub_scores = []
    for sid, s in score_map.items():
        pub_scores.append({k: v for k, v in s.items() if k != "needs_review"})

    # ---------- aliases.json ------------------------------------------------
    aliases = {
        "by_lookup": {},
    }
    for o in orgs:
        aliases["by_lookup"][o["_nk_name"]] = o["id"]

    # ---------- methodology.json --------------------------------------------
    methodology = {
        "version": METHODOLOGY_VERSION,
        "status": "beta",
        "formula": {
            "expression": "DeFeX = clamp(0, 100,  B - sum(Pi) + E)",
            "base_scores": BASE_SCORE,
            "penalties": {
                "vacancy_based": PENALTY_VACANCY,
                "senior_only": PENALTY_SENIOR_ONLY,
                "fcfs": PENALTY_FCFS,
                "highly_beneficial": 0,
            },
            "evidence_bonus": {"with_om": 8, "with_personal": 2, "max": 8},
            "caps": [
                "If any penalty applies, DeFeX is capped at 75.",
                "Missing data does not lower DeFeX — it lowers confidence."
            ],
        },
        "bands": [
            {"min": 80, "max": 100, "label": "Deputation-friendly"},
            {"min": 65, "max": 79,  "label": "Generally supportive"},
            {"min": 50, "max": 64,  "label": "Mixed"},
            {"min": 35, "max": 49,  "label": "Restrictive"},
            {"min": 0,  "max": 34,  "label": "Very restrictive"},
            {"label": "Unrated", "note": "Insufficient data — appears in Mapped tab only."}
        ],
        "confidence": {
            "weights": {
                "reports": 0.25, "official_source": 0.40, "recency": 0.15,
                "consistency": 0.10, "stage_coverage": 0.10
            },
            "bands": {"High": 0.70, "Medium": 0.45, "Low": 0.20}
        },
        "moderation": {
            "officer_names_policy": "Submissions are scanned for officer names and rejected — never redacted.",
            "integrity_flag_policy": "Submissions alleging integrity issues are NEVER displayed publicly. They route to manual review and may lower confidence."
        },
        "changelog": [
            {"version": "0.2.0-beta", "date": "2026-06-15",
             "note": "Cumulative scoring. An organisation's DeFeX now reflects ALL responses collected for it, not a single row: the base uses the consensus Deputation_Allowed, friction penalties apply only on a majority report, community reports add +2 each (capped +8), and confidence rises with response volume and agreement. Rankings are rebuilt monthly."},
            {"version": "0.1.0-beta", "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
             "note": "Initial public beta. Community survey was floated in February 2025 and received 114 responses; this version derives the DeFeX formula from those one-row-per-org policy snapshots. Per-applicant timeline reports (forwarding, NOC, vigilance, relieving) will begin populating as the survey continues to receive submissions."}
        ]
    }

    # ---------- updates.json ------------------------------------------------
    rated_orgs = [o for o in pub_orgs if o["rated"]]
    updates = {
        "generated_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "methodology_version": METHODOLOGY_VERSION,
        "source": source_label,
        "counts": {
            "organisations_mapped": len(pub_orgs),
            "ministries_covered": len({o["ministry"] for o in pub_orgs}),
            "organisations_rated": len(rated_orgs),
            "organisations_with_om": sum(1 for o in pub_orgs if o["has_om"]),
            "reports_total": len(reports),
            "unresolved_input_rows": len(unresolved),
        },
        "survey_url": SURVEY_URL,
    }

    # ---------- write -------------------------------------------------------
    (OUT_DIR / "organisations.json").write_text(json.dumps(pub_orgs, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / "reports.json").write_text(json.dumps(reports, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / "scores.json").write_text(json.dumps(pub_scores, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / "aliases.json").write_text(json.dumps(aliases, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / "methodology.json").write_text(json.dumps(methodology, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / "updates.json").write_text(json.dumps(updates, ensure_ascii=False, indent=2), encoding="utf-8")
    if unresolved:
        (OUT_DIR / "_unresolved.log.json").write_text(json.dumps(unresolved, ensure_ascii=False, indent=2), encoding="utf-8")

    # Public-safety assertion: Suspect_Bribe must not appear in any public JSON.
    for fname in ("organisations.json", "reports.json", "scores.json", "methodology.json", "updates.json", "aliases.json"):
        body = (OUT_DIR / fname).read_text(encoding="utf-8")
        assert "Suspect_Bribe" not in body and "suspect_bribe" not in body, f"LEAK in {fname}"

    print(f"source:         {source_label}")
    print(f"organisations: {len(pub_orgs)}")
    print(f"  rated:        {len(rated_orgs)}")
    print(f"  with OM:      {sum(1 for o in pub_orgs if o['has_om'])}")
    print(f"  ministries:   {updates['counts']['ministries_covered']}")
    print(f"reports:        {len(reports)}")
    print(f"unresolved:     {len(unresolved)} (see _unresolved.log.json)")


if __name__ == "__main__":
    main()
